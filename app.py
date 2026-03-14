import streamlit as st
import pandas as pd
import io
import os
import math
import re

# إعدادات الصفحة
st.set_page_config(page_title="توزيع أماكن الامتحانات", layout="wide")

# ==========================================
# 1. فلتر توحيد النصوص العربية (لمعالجة اختلافات الكتابة والهمزات)
# ==========================================
def normalize_arabic(text):
    if not isinstance(text, str):
        return ""
    t = text.strip()
    t = re.sub(r'[إأآا]', 'ا', t) # توحيد الألف
    t = re.sub(r'ة', 'ه', t)      # توحيد التاء المربوطة والهاء
    t = re.sub(r'[يى]', 'ى', t)   # توحيد الياء
    t = re.sub(r'\s+', ' ', t)    # إزالة المسافات الزايدة
    return t

# ==========================================
# 2. محرك الذكاء اللغوي لتفكيك ودمج الملاحظات
# ==========================================
def parse_level_string(raw_str):
    s = str(raw_str).replace("المستوي", "").replace("المستوى", "").strip()
    
    lvl = ""
    level_map = {'1': 'الاول', '2': 'الثاني', '3': 'الثالث', '4': 'الرابع', 
                 'الاول': 'الاول', 'الثاني': 'الثاني', 'الثالث': 'الثالث', 'الرابع': 'الرابع'}
    words = s.split()
    for k in level_map:
        if k in words:
            lvl = level_map[k]
            words.remove(k)
            s = " ".join(words)
            break
            
    mjr = ""
    majors = ["إدارة الأعمال", "ادارة الاعمال", "إداره الاعمال", "اداره الاعمال", 
              "الموارد البشرية", "موارد بشرية", "الموارد البشریة", "موارد بشریة", "موارد",
              "نظم المعلومات", "نظم معلومات", "المحاسبة", "محاسبة", "محاسبه", 
              "الإدارة", "الادارة", "إدارة", "ادارة", "إداره", "اداره", 
              "الإحصاء", "الاحصاء", "إحصاء", "احصاء", "التمويل", "تمويل", 
              "الجمارك", "جمارك", "التسويق", "تسويق", "النظم", "نظم"]
    majors.sort(key=len, reverse=True) 
    for m in majors:
        if m in s:
            mjr = m
            s = s.replace(m, "", 1).strip()
            break
            
    typ = ""
    for t in ["انتظام", "انتساب"]:
        if t in s:
            typ = t
            s = s.replace(t, "", 1).strip()
            break
            
    mod = " ".join(s.split())
    return lvl, mjr, typ, mod

def generate_smart_notes(raw_levels_set):
    if not raw_levels_set:
        return ""

    grouped_data = {}
    for raw in raw_levels_set:
        lvl, mjr, typ, mod = parse_level_string(raw)
        key = (lvl, typ)
        if key not in grouped_data:
            grouped_data[key] = {}
        if mjr not in grouped_data[key]:
            grouped_data[key][mjr] = set()
        if mod:
            grouped_data[key][mjr].add(mod)
        else:
            grouped_data[key][mjr].add("")

    level_order = {'الاول': 1, 'الثاني': 2, 'الثالث': 3, 'الرابع': 4}
    sorted_keys = sorted(list(grouped_data.keys()), key=lambda k: (level_order.get(k[0], 99), k[0], k[1]))
    results = []
    
    for lvl, typ in sorted_keys:
        majors_dict = grouped_data[(lvl, typ)]
        
        def sort_mods(mods_set):
            m_list = [m for m in set(mods_set) if m]
            m_list.sort() 
            return " و".join(m_list)

        if len(majors_dict) == 1:
            mjr = list(majors_dict.keys())[0]
            mods = majors_dict[mjr]
            mods_str = sort_mods(mods)
            parts = [p for p in [lvl, mjr, typ, mods_str] if p]
            results.append(" ".join(parts))
        else:
            all_mods = set()
            for m_mods in majors_dict.values():
                all_mods.update(m_mods)
            mods_str = sort_mods(all_mods)
            parts = [p for p in [lvl, typ, mods_str] if p]
            results.append(" ".join(parts))
            
    if results:
        return "المستوي " + " & ".join(results)
    return ""

# ==========================================
# ستايل الواجهة الأساسي
# ==========================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;700&display=swap');
    * { font-family: 'Tajawal', sans-serif !important; }
    .stApp, .stMarkdown, .stText, p, label, input, .stSelectbox { 
        text-align: right !important; direction: rtl !important; font-size: 16px !important;
    }
    h1 { text-align: center !important; color: #1E3A8A !important; font-weight: 700 !important; }
    h3 { 
        color: #004d40 !important; font-weight: 700 !important; padding-top: 15px; 
        border-bottom: 2px solid #b6e3f4; padding-bottom: 10px; margin-bottom: 20px; 
        text-align: right !important; direction: rtl !important;
    }
    .stButton > button { font-size: 18px !important; font-weight: bold !important; width: 100% !important; }
    
    [data-testid="stDataFrame"] div[role="gridcell"], 
    [data-testid="stDataFrame"] div[role="columnheader"] {
        text-align: right !important; justify-content: flex-end !important; font-size: 15px !important;
    }
    
    .stTabs [data-baseweb="tab-list"] { gap: 10px; justify-content: flex-end; }
    .stTabs [data-baseweb="tab"] { font-size: 18px; font-weight: bold; background-color: #f0f4f8; border-radius: 5px 5px 0 0; padding: 10px 20px; }
    .stTabs [aria-selected="true"] { background-color: #1E3A8A !important; color: white !important; }
    </style>
    """,
    unsafe_allow_html=True
)

# --- الهيدر ---
col_left, col_space, col_right = st.columns([1, 3, 1])
with col_left:
    if os.path.exists("logo_faculty.png"): st.image("logo_faculty.png", use_container_width=True)
    elif os.path.exists("logo_faculty.jpg"): st.image("logo_faculty.jpg", use_container_width=True)
with col_space:
    st.markdown("<div style='display: flex; justify-content: center; align-items: center; height: 100%; margin-top: 20px;'><h1 style='margin: 0;'>توزيع أماكن الامتحانات (الشجرة)</h1></div>", unsafe_allow_html=True)
with col_right:
    if os.path.exists("logo_unit.png"): st.image("logo_unit.png", use_container_width=True)
    elif os.path.exists("logo_unit.jpg"): st.image("logo_unit.jpg", use_container_width=True)

st.markdown("---")

if 'rooms_df' not in st.session_state: st.session_state.rooms_df = None
if 'students_df' not in st.session_state: st.session_state.students_df = None
if 'courses_order' not in st.session_state: st.session_state.courses_order = []

# ==========================================
# الخطوة 1: رفع الملفات
# ==========================================
if st.session_state.rooms_df is None or st.session_state.students_df is None:
    st.markdown("<h3>الخطوة 1: رفع ملفات البيانات الأساسية</h3>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)

    with col1: 
        st.markdown("**1. ملف أماكن اللجان (رقم اللجنة، مكان اللجنة، سعة اللجنة)**")
        rooms_file = st.file_uploader("ارفع ملف اللجان بصيغة Excel", type=['xlsx'], key="rooms_uploader")
        if rooms_file:
            try:
                df_rooms = pd.read_excel(rooms_file)
                df_rooms.columns = df_rooms.columns.str.strip()
                required_rooms = ["رقم اللجنة", "مكان اللجنة", "سعة اللجنة"]
                if all(col in df_rooms.columns for col in required_rooms):
                    st.session_state.rooms_df = df_rooms[required_rooms].copy()
                    st.success(f"✅ تم قراءة ملف اللجان بنجاح! (عدد اللجان: {len(df_rooms)})")
                else:
                    st.error("❌ تأكد من وجود الأعمدة: رقم اللجنة، مكان اللجنة، سعة اللجنة")
            except Exception as e:
                st.error(f"حدث خطأ أثناء قراءة ملف اللجان: {e}")

    with col2: 
        st.markdown("**2. ملف الطلبة (شيت واحد مجمع أو عدة شيتات)**")
        students_file = st.file_uploader("ارفع ملف الطلبة بصيغة Excel", type=['xlsx'], key="students_uploader")
        if students_file:
            try:
                all_sheets = pd.read_excel(students_file, sheet_name=None)
                all_students = []
                extracted_courses_order = []
                
                for sheet_name, df in all_sheets.items():
                    df.columns = df.columns.str.strip()
                    
                    col_mapping = {
                        "المستوى": "المستوي",
                        "المقرر": "اسم المقرر",
                        "اسم المادة": "اسم المقرر",
                        "الماده": "اسم المقرر",
                        "رقم جلوس": "رقم الجلوس",
                        "ترتيب المواد": "ترتيب المقررات"
                    }
                    df.rename(columns=col_mapping, inplace=True)
                    
                    if "ترتيب المقررات" in df.columns:
                        order_list = df["ترتيب المقررات"].dropna().astype(str).str.strip().tolist()
                        for c in order_list:
                            if c and c != "nan" and c not in extracted_courses_order:
                                extracted_courses_order.append(c)
                    
                    required_students = ["رقم الجلوس", "اسم المقرر", "المستوي"]
                    if all(col in df.columns for col in required_students):
                        all_students.append(df[required_students])
                    else:
                        st.warning(f"⚠️ الشيت '{sheet_name}' تم تجاهله لعدم وجود الأعمدة المطلوبة.")
                
                st.session_state.courses_order = extracted_courses_order

                if all_students:
                    df_all = pd.concat(all_students, ignore_index=True)
                    df_all['رقم الجلوس'] = pd.to_numeric(df_all['رقم الجلوس'], errors='coerce')
                    df_all.dropna(subset=['رقم الجلوس'], inplace=True)
                    df_all['رقم الجلوس'] = df_all['رقم الجلوس'].astype(int)
                    st.session_state.students_df = df_all
                    st.success(f"✅ تم دمج وقراءة بيانات الطلبة بنجاح! (إجمالي السجلات: {len(df_all)})")
                    st.rerun()
                else:
                    st.error("❌ لم يتم العثور على الأعمدة المطلوبة في أي شيت.")
            except Exception as e:
                st.error(f"حدث خطأ أثناء قراءة ملف الطلبة: {e}")

# ==========================================
# الخطوة 2 و 3: إدخال البيانات والخوارزمية
# ==========================================
if st.session_state.rooms_df is not None and st.session_state.students_df is not None:
    
    # --- الخطوة 2 ---
    st.markdown("<h3>الخطوة 2: إدخال بيانات الخريطة</h3>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        exam_period = st.selectbox("أماكن امتحانات:", ["منتصف فصل الخريف", "نهاية فصل الخريف", "منتصف فصل الربيع", "نهاية فصل الربيع"])
    with c2:
        academic_year = st.selectbox("العام الجامعي:", ["2025 - 2026", "2026 - 2027", "2027 - 2028", "2028 - 2029", "2029 - 2030"])
    
    level_courses = st.text_input("مقررات المستوي:")
    
    # --- الخطوة 3 ---
    st.markdown("<h3>الخطوة 3: توليد خريطة اللجان الموحدة</h3>", unsafe_allow_html=True)
    
    df_students = st.session_state.students_df
    unique_seats = sorted(df_students['رقم الجلوس'].unique())
    total_unique_students = len(unique_seats)
    
    # ==========================================
    # ترتيب المقررات مع التغاضي عن الأخطاء الإملائية
    # ==========================================
    raw_subjects = df_students['اسم المقرر'].unique()
    all_subjects = []
    
    if st.session_state.courses_order:
        for ordered_c in st.session_state.courses_order:
            norm_ordered = normalize_arabic(ordered_c)
            # البحث عن المادة في البيانات الفعلية بعد توحيد النص
            for actual_c in raw_subjects:
                if normalize_arabic(actual_c) == norm_ordered and actual_c not in all_subjects:
                    all_subjects.append(actual_c)
                
    remaining_subjects = sorted([c for c in raw_subjects if c not in all_subjects])
    all_subjects.extend(remaining_subjects)
    # ==========================================
    
    seat_courses = df_students.groupby('رقم الجلوس')['اسم المقرر'].apply(lambda x: list(set(x))).to_dict()
    seat_levels = df_students.groupby('رقم الجلوس')['المستوي'].apply(lambda x: list(set(x))).to_dict()
    
    st.info(f"إجمالي عدد الطلبة (بدون تكرار) المطلوب توزيعهم: **{total_unique_students}** طالب.")
    if st.session_state.courses_order:
        st.success("✅ تم تفعيل ترتيب المقررات المخصص وتوحيد الأسماء بنجاح.")
    
    if st.button("🚀 بدء التوزيع وتوليد الوثائق الرسمية", type="primary"):
        with st.spinner("جاري التوزيع وتطبيق معالجة الملاحظات الدقيقة (سعة صارمة بدون +1)..."):
            result_data = []
            curr_student_idx = 0
            
            first_seat = unique_seats[0] if total_unique_students > 0 else 0
            current_range_start = math.floor(first_seat / 5.0) * 5 if first_seat > 0 else 0
            if current_range_start < first_seat and current_range_start % 10 == 0:
                current_range_start += 1
            
            rooms_list = st.session_state.rooms_df.to_dict('records')
            
            for room in rooms_list:
                room_num = room['رقم اللجنة']
                room_loc = room['مكان اللجنة']
                try: room_cap = int(room['سعة اللجنة'])
                except: room_cap = 0
                
                if curr_student_idx >= total_unique_students or room_cap <= 0:
                    empty_room = {
                        'رقم اللجنة': room_num, 'مكان اللجنة': room_loc, 'سعة اللجنة': room_cap,
                        'من': '-', 'إلى': '-', 'ملاحظات': 'لجنة فارغة'
                    }
                    for subj in all_subjects: empty_room[subj] = '-'
                    result_data.append(empty_room)
                    continue
                
                course_counts = {}
                max_possible_c = 0
                for i in range(curr_student_idx, total_unique_students):
                    seat = unique_seats[i]
                    courses_for_seat = seat_courses.get(seat, [])
                    
                    can_add = True
                    for c in courses_for_seat:
                        if course_counts.get(c, 0) + 1 > room_cap: 
                            can_add = False
                            break
                            
                    remaining_total_students = total_unique_students - i
                    if not can_add and remaining_total_students <= 4:
                        can_add = True 
                    
                    if not can_add:
                        break 
                        
                    for c in courses_for_seat:
                        course_counts[c] = course_counts.get(c, 0) + 1
                    max_possible_c += 1
                
                final_end = None
                best_c = max_possible_c
                
                if curr_student_idx + max_possible_c == total_unique_students:
                    best_c = max_possible_c
                    last_actual = unique_seats[curr_student_idx + best_c - 1]
                    final_end = math.ceil(last_actual / 5.0) * 5
                else:
                    found_nice_end = False
                    for test_c in range(max_possible_c, 0, -1):
                        temp_counts = {}
                        for i in range(test_c):
                            for c in seat_courses.get(unique_seats[curr_student_idx + i], []):
                                temp_counts[c] = temp_counts.get(c, 0) + 1
                        current_max_load = max(temp_counts.values()) if temp_counts else 0
                        
                        if current_max_load < room_cap - 3:
                            break
                            
                        last_included = unique_seats[curr_student_idx + test_c - 1]
                        next_actual = unique_seats[curr_student_idx + test_c]
                        largest_multiple_of_5 = math.floor((next_actual - 1) / 5.0) * 5
                        
                        if largest_multiple_of_5 >= last_included:
                            final_end = largest_multiple_of_5
                            best_c = test_c
                            found_nice_end = True
                            break
                            
                    if not found_nice_end:
                        target_c = max_possible_c
                        for test_c in range(max_possible_c, 0, -1):
                            temp_counts = {}
                            for i in range(test_c):
                                for c in seat_courses.get(unique_seats[curr_student_idx + i], []):
                                    temp_counts[c] = temp_counts.get(c, 0) + 1
                            current_max_load = max(temp_counts.values()) if temp_counts else 0
                            if current_max_load <= room_cap:
                                target_c = test_c
                                break
                                
                        best_c = target_c
                        next_actual = unique_seats[curr_student_idx + best_c]
                        final_end = next_actual - 1
                
                final_course_counts = {}
                room_levels = set()
                
                for i in range(curr_student_idx, curr_student_idx + best_c):
                    current_seat = unique_seats[i]
                    for c in seat_courses.get(current_seat, []):
                         final_course_counts[c] = final_course_counts.get(c, 0) + 1
                    
                    for lvl in seat_levels.get(current_seat, []):
                        room_levels.add(str(lvl))
                
                notes_text = generate_smart_notes(room_levels)
                
                room_data = {
                    'رقم اللجنة': room_num,
                    'مكان اللجنة': room_loc,
                    'سعة اللجنة': room_cap,
                    'من': current_range_start,
                    'إلى': final_end,
                    'ملاحظات': notes_text
                }
                for subj in all_subjects:
                    room_data[subj] = final_course_counts.get(subj, 0)
                    
                result_data.append(room_data)
                
                current_range_start = final_end + 1
                curr_student_idx += best_c
            
            final_df = pd.DataFrame(result_data)
            
            summary_data = []
            for row in result_data:
                summary_data.append({
                    'رقم اللجنة': row['رقم اللجنة'],
                    'مكان اللجنة': row['مكان اللجنة'],
                    'بداية اللجنة (من)': row['من'],
                    'نهاية اللجنة (إلى)': row['إلى'],
                    'ملاحظات': row['ملاحظات']
                })
            summary_df = pd.DataFrame(summary_data)

            if curr_student_idx < total_unique_students:
                st.error(f"⚠️ تحذير: إجمالي سعة اللجان لا تكفي! متبقي {total_unique_students - curr_student_idx} طالب بدون أماكن.")
            else:
                st.success("✅ تم الانتهاء من التوزيع وإنشاء الشيتين بنجاح!")

            tab1, tab2 = st.tabs(["📄 خريطة اللجان (ملخص)", "📊 الخريطة التفصيلية (بالمواد)"])
            with tab1:
                display_summary_cols = summary_df.columns.tolist()[::-1]
                styled_summary = summary_df[display_summary_cols].style.set_properties(**{'text-align': 'right'}).set_table_styles([dict(selector='th', props=[('text-align', 'right')])])
                st.dataframe(styled_summary, hide_index=True, use_container_width=True)
            with tab2:
                display_final_cols = final_df.columns.tolist()[::-1]
                styled_final = final_df[display_final_cols].style.set_properties(**{'text-align': 'right'}).set_table_styles([dict(selector='th', props=[('text-align', 'right')])])
                st.dataframe(styled_final, hide_index=True, use_container_width=True)
            
            # --- توليد ملف الإكسيل الاحترافي ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                
                sheets_info = [
                    {'df': summary_df, 'name': 'خريطة اللجان', 'orientation': 'portrait'},
                    {'df': final_df, 'name': 'الخريطة التفصيلية', 'orientation': 'landscape'}
                ]
                
                from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
                from openpyxl.worksheet.table import Table, TableStyleInfo
                from openpyxl.utils import get_column_letter
                from openpyxl.drawing.image import Image as xlImage
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center', readingOrder=2)
                right_align = Alignment(horizontal='right', vertical='center', readingOrder=2)
                
                empty_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                header_font_white = Font(color="FFFFFF", bold=True, size=12)
                data_font = Font(bold=True, size=12) 
                meta_font = Font(bold=True, size=16, color="1E3A8A") 
                
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                red_bold_font = Font(color="FF0000", bold=True, size=12)
                
                fac_logo = "logo_faculty.png" if os.path.exists("logo_faculty.png") else "logo_faculty.jpg" if os.path.exists("logo_faculty.jpg") else None
                unit_logo = "logo_unit.png" if os.path.exists("logo_unit.png") else "logo_unit.jpg" if os.path.exists("logo_unit.jpg") else None
                
                for idx, info in enumerate(sheets_info):
                    current_df = info['df']
                    sheet_name = info['name']
                    orientation = info['orientation']
                    
                    total_columns = len(current_df.columns)
                    last_col_letter = get_column_letter(total_columns)
                    
                    current_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=5)
                    worksheet = writer.sheets[sheet_name]
                    worksheet.sheet_view.rightToLeft = True 
                    
                    target_h = 100 
                    try:
                        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                        from openpyxl.drawing.xdr import XDRPositiveSize2D
                        from openpyxl.utils.units import pixels_to_EMU
                        
                        if fac_logo:
                            img1 = xlImage(fac_logo)
                            ratio1 = target_h / img1.height
                            img1.width = int(img1.width * ratio1)
                            img1.height = int(target_h)
                            
                            size1 = XDRPositiveSize2D(cx=pixels_to_EMU(img1.width), cy=pixels_to_EMU(img1.height))
                            marker1 = AnchorMarker(col=0, colOff=pixels_to_EMU(10), row=0, rowOff=pixels_to_EMU(5))
                            img1.anchor = OneCellAnchor(_from=marker1, ext=size1)
                            worksheet.add_image(img1)
                            
                        if unit_logo:
                            img2 = xlImage(unit_logo)
                            ratio2 = target_h / img2.height
                            img2.width = int(img2.width * ratio2)
                            img2.height = int(target_h)
                            
                            col_idx = total_columns - 1
                            if sheet_name == 'خريطة اللجان':
                                col_w = 50
                            else:
                                col_w = 16 if total_columns > 6 else 45
                                
                            col_w_px = col_w * 7.5 
                            offset_px = int(col_w_px - img2.width - 10) 
                            if offset_px < 0: offset_px = 0
                            
                            size2 = XDRPositiveSize2D(cx=pixels_to_EMU(img2.width), cy=pixels_to_EMU(img2.height))
                            marker2 = AnchorMarker(col=col_idx, colOff=pixels_to_EMU(offset_px), row=0, rowOff=pixels_to_EMU(5))
                            img2.anchor = OneCellAnchor(_from=marker2, ext=size2)
                            worksheet.add_image(img2)
                            
                    except Exception:
                        if fac_logo:
                            img1 = xlImage(fac_logo)
                            ratio1 = target_h / img1.height
                            img1.width, img1.height = int(img1.width * ratio1), int(target_h)
                            worksheet.add_image(img1, 'A1')
                        if unit_logo:
                            img2 = xlImage(unit_logo)
                            ratio2 = target_h / img2.height
                            img2.width, img2.height = int(img2.width * ratio2), int(target_h)
                            worksheet.add_image(img2, f'{last_col_letter}1')
                    
                    if total_columns > 2:
                        merge_start = 'B'
                        merge_end = get_column_letter(total_columns - 1)
                    else:
                        merge_start = 'A'
                        merge_end = 'A'
                        
                    if merge_start != merge_end:
                        worksheet.merge_cells(f'{merge_start}1:{merge_end}1')
                        worksheet.merge_cells(f'{merge_start}2:{merge_end}2')
                        worksheet.merge_cells(f'{merge_start}3:{merge_end}3')
                        
                    worksheet[f'{merge_start}1'] = f"أماكن امتحانات: {exam_period}"
                    worksheet[f'{merge_start}2'] = f"العام الجامعي: {academic_year}"
                    worksheet[f'{merge_start}3'] = f"مقررات المستوي: {level_courses}"
                    
                    for r in range(1, 6):
                        worksheet.row_dimensions[r].height = 35 
                        if r <= 3:
                            cell = worksheet[f'{merge_start}{r}']
                            cell.alignment = center_align
                            cell.font = meta_font
                    
                    last_row = worksheet.max_row
                    
                    table_ref = f"A6:{last_col_letter}{last_row}"
                    tab = Table(displayName=f"TableMap_{idx}", ref=table_ref)
                    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    tab.tableStyleInfo = style
                    worksheet.add_table(tab)
                    
                    for r_idx in range(6, last_row + 1):
                        worksheet.row_dimensions[r_idx].height = 26.25 
                        
                        is_empty = False
                        capacity_val = 0
                        
                        if r_idx > 6:
                            if sheet_name == 'خريطة اللجان':
                                is_empty = (worksheet.cell(row=r_idx, column=3).value == '-') 
                            else:
                                is_empty = (worksheet.cell(row=r_idx, column=4).value == '-') 
                                try:
                                    capacity_val = int(worksheet.cell(row=r_idx, column=3).value)
                                except:
                                    capacity_val = 0
                                
                        for c_idx in range(1, total_columns + 1):
                            cell = worksheet.cell(row=r_idx, column=c_idx)
                            cell.border = thin_border
                            
                            if r_idx == 6: 
                                cell.font = header_font_white
                                cell.alignment = center_align
                            else:
                                cell.font = data_font 
                                
                                if c_idx == 2: 
                                    cell.alignment = right_align
                                else:
                                    cell.alignment = center_align
                                    
                                if is_empty:
                                    cell.fill = empty_fill
                                else:
                                    if sheet_name == 'الخريطة التفصيلية' and c_idx > 6:
                                        try:
                                            subject_count = int(cell.value)
                                            if subject_count > capacity_val:
                                                cell.fill = yellow_fill
                                                cell.font = red_bold_font
                                        except:
                                            pass
                                
                    if sheet_name == 'خريطة اللجان':
                        worksheet.column_dimensions['A'].width = 15 
                        worksheet.column_dimensions['B'].width = 45 
                        worksheet.column_dimensions['C'].width = 20 
                        worksheet.column_dimensions['D'].width = 20 
                        worksheet.column_dimensions['E'].width = 50 
                    else:
                        worksheet.column_dimensions['A'].width = 15 
                        worksheet.column_dimensions['B'].width = 35 
                        worksheet.column_dimensions['C'].width = 12 
                        worksheet.column_dimensions['D'].width = 15 
                        worksheet.column_dimensions['E'].width = 15 
                        worksheet.column_dimensions['F'].width = 45 
                        for i in range(7, total_columns + 1):
                            worksheet.column_dimensions[get_column_letter(i)].width = 16
                            
                    worksheet.print_area = f"A1:{last_col_letter}{last_row}"
                    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
                    
                    if orientation == 'landscape':
                        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
                    else:
                        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
                        
                    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                    worksheet.page_setup.fitToWidth = 1
                    worksheet.page_setup.fitToHeight = 0
                    worksheet.print_options.horizontalCentered = True
                    
                    worksheet.print_title_rows = '1:6'
                    worksheet.oddFooter.center.text = "&12 صفحة رقم (&P) من (&N)"
                    worksheet.evenFooter.center.text = "&12 صفحة رقم (&P) من (&N)"

            st.markdown("<div style='display: flex; justify-content: flex-end; width: 100%; margin-top: 15px;'>", unsafe_allow_html=True)
            
            safe_exam = exam_period.replace(" ", "_")
            safe_year = academic_year.replace(" ", "")
            if level_courses.strip():
                level_part = f"المستوي_{level_courses.strip()}"
            else:
                level_part = "غير_محدد"
                
            dynamic_file_name = f"خريطة_لجان_{level_part}_{safe_exam}_{safe_year}.xlsx"
            
            st.download_button(
                label="📥 تحميل خريطة اللجان الرسمية (Excel)", 
                data=output.getvalue(), 
                file_name=dynamic_file_name, 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                type="primary"
            )
            st.markdown("</div>", unsafe_allow_html=True)
            
    st.markdown("---")
    if st.button("تفريغ البيانات لرفع ملف جديدة"):
        st.session_state.rooms_df = None
        st.session_state.students_df = None
        st.session_state.courses_order = []
        st.rerun()
